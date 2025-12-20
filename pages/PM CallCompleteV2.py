import streamlit as st
import pandas as pd
from io import BytesIO
import barcode
from barcode.writer import ImageWriter
from PIL import Image
import base64
from streamlit_qrcode_scanner import qrcode_scanner

# =====================
# CONFIG
# =====================
st.set_page_config(page_title="PM SST – Guided Scan", layout="centered")
st.title("🧾 PM SST – Guided Component Scan")
st.caption("Follow the exact component order shown inside the machine")
st.divider()

# =====================
# SEQUÊNCIA OFICIAL (15 componentes)
# =====================
PM_SEQUENCE = [
    "Burster 1",
    "Burster 2",
    "Burster 3",
    "Burster 4",
    "Burster 5",
    "Burster 6",
    "Burster 7",
    "Scanner",
    "Printer",
    "Slip Reader",
    "LCD",
    "Keypad",
    "Enclosure",
    "Router",
    "Pin Pad",
]

TOTAL_STEPS = len(PM_SEQUENCE)

# =====================
# FUNÇÃO PARA GERAR CODE 128
# =====================
def generate_barcode(serial_number):
    """Gera código de barras Code 128 e retorna como bytes PNG"""
    try:
        # Cria código Code 128
        code128 = barcode.get_barcode_class('code128')
        
        # Gera o código de barras
        barcode_obj = code128(serial_number, writer=ImageWriter())
        
        # Configurações do escritor para melhor legibilidade
        writer_options = {
            'module_height': 10.0,  # Altura das barras
            'module_width': 0.3,    # Largura das barras
            'quiet_zone': 6.0,      # Zona silenciosa
            'font_size': 12,        # Tamanho da fonte do texto
            'text_distance': 4.0,   # Distância do texto às barras
            'write_text': True,     # Mostrar texto abaixo
            'background': 'white',  # Fundo branco
            'foreground': 'black',  # Barras pretas
        }
        
        # Salva em bytes
        barcode_bytes = BytesIO()
        barcode_obj.write(barcode_bytes, options=writer_options)
        barcode_bytes.seek(0)
        
        return barcode_bytes.getvalue()
        
    except Exception as e:
        st.error(f"Error generating barcode: {e}")
        return None

# =====================
# FUNÇÃO PARA BASE64 (para visualização)
# =====================
def get_barcode_base64(barcode_bytes):
    """Converte bytes da imagem para base64"""
    if barcode_bytes:
        return base64.b64encode(barcode_bytes).decode()
    return None

# =====================
# SESSION STATE INIT
# =====================
if "pm_step" not in st.session_state:
    st.session_state.pm_step = 0

if "pm_data" not in st.session_state:
    st.session_state.pm_data = {
        comp: {
            "Serial": "", 
            "Barcode": "", 
            "Status": "PENDING",
            "BarcodeImage": None,  # Para armazenar a imagem gerada
            "BarcodeBase64": None  # Para visualização
        }
        for comp in PM_SEQUENCE
    }

if "pm_scanned" not in st.session_state:
    st.session_state.pm_scanned = {
        comp: False for comp in PM_SEQUENCE
    }

if "scanner_buffer" not in st.session_state:
    st.session_state.scanner_buffer = None

if "scanned_codes" not in st.session_state:
    st.session_state.scanned_codes = []

# =====================
# PASSO ATUAL
# =====================
step = st.session_state.pm_step

if step < TOTAL_STEPS:
    component = PM_SEQUENCE[step]
    
    # Pula se já escaneado
    if st.session_state.pm_scanned[component]:
        st.session_state.pm_step += 1
        st.rerun()

    st.subheader(f"Step {step + 1} / {TOTAL_STEPS}")
    st.markdown(f"## 🔒 CURRENT COMPONENT: **{component}**")
    
    # Progresso
    progress = step / TOTAL_STEPS
    st.progress(progress)

    # =====================
    # VISUALIZAÇÃO DO CÓDIGO GERADO (se existir)
    # =====================
    current_data = st.session_state.pm_data[component]
    
    if current_data["BarcodeImage"] and current_data["BarcodeBase64"]:
        st.markdown("### 📊 Generated Barcode Preview")
        
        col1, col2 = st.columns([1, 2])
        
        with col1:
            # Mostra a imagem do código de barras
            st.image(
                current_data["BarcodeImage"],
                caption=f"Code 128: {current_data['Serial']}",
                use_column_width=True
            )
        
        with col2:
            # Informações
            st.info(f"""
            **Component:** {component}
            **Serial Number:** {current_data['Serial']}
            **Barcode:** {current_data['Barcode']}
            
            *This barcode will be included in the labels for printing.*
            """)

    # Campos de entrada
    col1, col2 = st.columns(2)
    
    with col1:
        st.text_input(
            "Serial Number",
            value=current_data["Serial"],
            disabled=True,
            key=f"serial_{component}"
        )
    
    with col2:
        st.text_input(
            "Barcode",
            value=current_data["Barcode"],
            disabled=True,
            key=f"barcode_{component}"
        )

    # =====================
    # SCANNER
    # =====================
    st.markdown("### 📷 Scan original barcode")
    st.info(f"Point camera at {component} original barcode")

    scanned = qrcode_scanner(key=f"scanner_{component}")

    if scanned and scanned != st.session_state.scanner_buffer:
        st.session_state.scanner_buffer = scanned
        
        # Verifica duplicação
        if scanned in st.session_state.scanned_codes:
            st.error(f"❌ This barcode was already scanned!")
            st.session_state.scanner_buffer = None
            st.rerun()
        
        # Validação
        if len(scanned) < 3:
            st.error("Invalid barcode. Please rescan.")
            st.session_state.scanner_buffer = None
        else:
            # GERA O NOVO CÓDIGO DE BARRAS!
            barcode_bytes = generate_barcode(scanned)
            
            if barcode_bytes:
                # Converte para base64 para visualização
                barcode_base64 = get_barcode_base64(barcode_bytes)
                
                # Atualiza dados
                st.session_state.pm_data[component]["Serial"] = scanned
                st.session_state.pm_data[component]["Barcode"] = scanned
                st.session_state.pm_data[component]["Status"] = "SCANNED"
                st.session_state.pm_data[component]["BarcodeImage"] = barcode_bytes
                st.session_state.pm_data[component]["BarcodeBase64"] = barcode_base64
                
                st.session_state.pm_scanned[component] = True
                st.session_state.scanned_codes.append(scanned)
                st.session_state.scanner_buffer = None
                
                # Feedback
                st.success(f"✅ {component} scanned and barcode generated!")
                st.balloons()
                
                # Mostra preview
                col1, col2 = st.columns(2)
                with col1:
                    st.image(barcode_bytes, caption="Generated Code 128", use_column_width=True)
                
                with col2:
                    st.code(f"""
                    Component: {component}
                    Serial: {scanned}
                    Barcode Type: Code 128
                    """)
                
                # Auto-advance
                import time
                time.sleep(2)
                st.session_state.pm_step += 1
                st.rerun()
            else:
                st.error("Failed to generate barcode. Please try again.")

    # =====================
    # ENTRADA MANUAL (fallback)
    # =====================
    st.divider()
    with st.expander("✏️ Manual Entry (if scanner fails)"):
        manual_serial = st.text_input(
            f"Enter serial number for {component}",
            key=f"manual_{component}"
        )
        
        if st.button(f"Generate Barcode for {component}", key=f"btn_manual_{component}"):
            if manual_serial and len(manual_serial) >= 3:
                # Gera código de barras
                barcode_bytes = generate_barcode(manual_serial)
                
                if barcode_bytes:
                    barcode_base64 = get_barcode_base64(barcode_bytes)
                    
                    # Atualiza dados
                    st.session_state.pm_data[component]["Serial"] = manual_serial
                    st.session_state.pm_data[component]["Barcode"] = manual_serial
                    st.session_state.pm_data[component]["Status"] = "SCANNED"
                    st.session_state.pm_data[component]["BarcodeImage"] = barcode_bytes
                    st.session_state.pm_data[component]["BarcodeBase64"] = barcode_base64
                    
                    st.session_state.pm_scanned[component] = True
                    if manual_serial not in st.session_state.scanned_codes:
                        st.session_state.scanned_codes.append(manual_serial)
                    
                    st.success(f"✅ Barcode generated for {component}!")
                    st.rerun()
            else:
                st.error("Please enter a valid serial number (min 3 characters)")

    # Controles de navegação
    st.divider()
    col1, col2 = st.columns(2)
    with col1:
        if st.button("⏮️ Previous", disabled=(step == 0)):
            st.session_state.pm_step -= 1
            st.rerun()
    
    with col2:
        if st.button("⏭️ Skip Component"):
            st.warning(f"Skipping {component}")
            st.session_state.pm_data[component]["Status"] = "SKIPPED"
            st.session_state.pm_scanned[component] = True
            st.session_state.pm_step += 1
            st.rerun()

else:
    # =====================
    # FINALIZAÇÃO - GERAR EXCEL COM CÓDIGOS DE BARRAS
    # =====================
    st.success("✅ PM Component Scan Completed!")
    st.balloons()
    
    st.subheader("📊 Final Report with Barcodes")
    
    # Prepara dados para DataFrame
    report_data = []
    barcode_images = []  # Para armazenar imagens
    
    for comp in PM_SEQUENCE:
        data = st.session_state.pm_data[comp]
        
        # Adiciona ao relatório
        report_data.append({
            "Component": comp,
            "Serial Number": data["Serial"] if data["Serial"] else "NOT SCANNED",
            "Barcode": data["Barcode"] if data["Barcode"] else "NOT SCANNED",
            "Status": data["Status"],
            "BarcodeImage": data["BarcodeImage"]  # Mantém os bytes da imagem
        })
        
        # Guarda referência da imagem
        if data["BarcodeImage"]:
            barcode_images.append((comp, data["BarcodeImage"]))
    
    df = pd.DataFrame(report_data)
    
    # =====================
    # VISUALIZAÇÃO DOS CÓDIGOS DE BARRAS
    # =====================
    with st.expander("👀 Preview All Generated Barcodes"):
        cols = st.columns(3)
        for idx, (comp, barcode_bytes) in enumerate(barcode_images):
            with cols[idx % 3]:
                st.image(barcode_bytes, caption=comp, use_column_width=True)
    
    # =====================
    # GERAR EXCEL COM CÓDIGOS DE BARRAS
    # =====================
    st.divider()
    st.subheader("📥 Export to Excel with Barcodes")
    
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    
    # Cria Excel
    output = BytesIO()
    wb = Workbook()
    
    # Sheet 1: Dados principais + imagens
    ws1 = wb.active
    ws1.title = "PM_SST_Labels"
    
    # Cabeçalhos
    headers = ["Component", "Serial Number", "Barcode", "Status", "Label Preview"]
    ws1.append(headers)
    
    # Adiciona dados e imagens
    for idx, row in df.iterrows():
        ws1.append([
            row["Component"],
            row["Serial Number"],
            row["Barcode"],
            row["Status"],
            f"See barcode image"
        ])
        
        # Se tem imagem, adiciona ao Excel
        if row["BarcodeImage"]:
            # Salva imagem temporariamente
            img = OpenpyxlImage(BytesIO(row["BarcodeImage"]))
            
            # Posiciona a imagem ao lado da linha
            cell_ref = f"E{idx + 2}"  # Coluna E, linha correspondente
            ws1.add_image(img, cell_ref)
            
            # Ajusta altura da linha para caber a imagem
            ws1.row_dimensions[idx + 2].height = 80
    
    # Ajusta largura das colunas
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 30
    
    # Sheet 2: Template de etiquetas (formato da foto)
    ws2 = wb.create_sheet("Print_Labels")
    
    # Cabeçalho do template
    ws2.append(["COMPONENT LABELS - FOR PRINTING"])
    ws2.append([])
    
    # Adiciona cada etiqueta no formato da foto
    for comp in PM_SEQUENCE:
        data = st.session_state.pm_data[comp]
        
        if data["Status"] == "SCANNED":
            # Formato igual à foto:
            # Nome do Equipamento
            # Código do código de barras (texto)
            # [Imagem do código]
            
            ws2.append([f"{comp}"])
            ws2.append([f"Serial: {data['Serial']}"])
            
            # Adiciona imagem se existir
            if data["BarcodeImage"]:
                img = OpenpyxlImage(BytesIO(data["BarcodeImage"]))
                img.width = 200
                img.height = 80
                ws2.add_image(img, f"A{ws2.max_row + 1}")
                
                # Pula algumas linhas para próxima etiqueta
                for _ in range(10):
                    ws2.append([])
    
    # Sheet 3: Resumo
    ws3 = wb.create_sheet("Summary")
    ws3.append(["PM SST - BARCODE GENERATION REPORT"])
    ws3.append([])
    ws3.append(["Scan Date:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")])
    ws3.append(["Total Components:", TOTAL_STEPS])
    ws3.append(["Scanned:", len([x for x in df["Status"] if x == "SCANNED"])])
    ws3.append(["Skipped:", len([x for x in df["Status"] if x == "SKIPPED"])])
    ws3.append([])
    ws3.append(["INSTRUCTIONS FOR PRINTING:"])
    ws3.append(["1. Print 'Print_Labels' sheet on label paper"])
    ws3.append(["2. Cut along dotted lines"])
    ws3.append(["3. Attach label to corresponding component"])
    ws3.append(["4. Keep this report for future reference"])
    
    # Salva Excel
    wb.save(output)
    output.seek(0)
    
    # Botão de download
    timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M')
    st.download_button(
        label=f"⬇️ Download Excel with Barcodes (.xlsx)",
        data=output.getvalue(),
        file_name=f"PM_SST_Barcodes_{timestamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # =====================
    # VISUALIZAÇÃO DAS ETIQUETAS
    # =====================
    st.divider()
    st.subheader("🏷️ Label Preview (Format for Printing)")
    
    # Mostra algumas etiquetas como exemplo
    sample_labels = []
    for comp in PM_SEQUENCE[:3]:  # Apenas 3 primeiros como exemplo
        data = st.session_state.pm_data[comp]
        if data["Status"] == "SCANNED" and data["BarcodeBase64"]:
            sample_labels.append({
                "component": comp,
                "serial": data["Serial"],
                "barcode_base64": data["BarcodeBase64"]
            })
    
    if sample_labels:
        cols = st.columns(len(sample_labels))
        for idx, label in enumerate(sample_labels):
            with cols[idx]:
                st.markdown(f"**{label['component']}**")
                st.markdown(f"`{label['serial']}`")
                st.image(
                    f"data:image/png;base64,{label['barcode_base64']}",
                    use_column_width=True
                )
                st.markdown("---")
    
    # =====================
    # RESET
    # =====================
    st.divider()
    if st.button("🔁 Start New PM Scan", type="primary"):
        keys_to_clear = [
            "pm_step", "pm_data", "pm_scanned", 
            "scanner_buffer", "scanned_codes"
        ]
        for key in keys_to_clear:
            if key in st.session_state:
                del st.session_state[key]
        st.rerun()
